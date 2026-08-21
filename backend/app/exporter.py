"""Export a register or the validation issues to CSV.

Sensitive fields (passwords, keys) are excluded from a plain export by default —
the register can hold a "Default Password" column and secret keys.
"""
from __future__ import annotations

import csv
import io
import json
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SENSITIVE = ("password", "private_key", "secret", "default_password")

# Report styling — the Electracom slate header used across the toolset.
_HEADER_FILL = PatternFill("solid", fgColor="26718F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(bold=True, size=14, color="2C2A28")


def _style_sheet(ws, headers: list[str], title: str | None = None) -> int:
    row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
        row = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    return row + 1


def _autofit(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        width = max((len(str(cell.value)) for cell in ws[letter] if cell.value), default=10)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 48)


def _is_sensitive(field_key: str) -> bool:
    fk = field_key.lower()
    return any(s in fk for s in SENSITIVE)


def export_register_csv(conn: sqlite3.Connection, project_id: int,
                        trade_id: int | None = None,
                        include_sensitive: bool = False) -> str:
    fields = [dict(r) for r in conn.execute(
        "SELECT field_key, display_name FROM schema_field "
        "WHERE project_id=? AND visible=1 AND field_key!='instance_name' "
        "ORDER BY export_order", (project_id,))]
    if not include_sensitive:
        fields = [f for f in fields if not _is_sensitive(f["field_key"])]
    cols = ["Trade", "Instance Name"] + [f["display_name"] for f in fields]

    q = ("SELECT a.instance_name, a.metadata, t.code AS trade_code FROM asset a "
         "JOIN trade t ON t.id=a.trade_id WHERE a.project_id=?")
    params: list = [project_id]
    if trade_id:
        q += " AND a.trade_id=?"
        params.append(trade_id)
    q += " ORDER BY t.code, a.instance_name"

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(cols)
    for a in conn.execute(q, params):
        meta = json.loads(a["metadata"] or "{}")
        row = [a["trade_code"], a["instance_name"]]
        row += [meta.get(f["field_key"], "") for f in fields]
        w.writerow(row)
    return buf.getvalue()


def export_issues_csv(conn: sqlite3.Connection, project_id: int) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Severity", "Rule", "Trade", "Instance Name", "Field", "Message", "Expected"])
    rows = conn.execute(
        "SELECT vi.severity, vi.rule, vi.field_key, vi.message, vi.expected, "
        "a.instance_name, t.code AS trade_code "
        "FROM validation_issue vi "
        "LEFT JOIN asset a ON a.id=vi.asset_id "
        "LEFT JOIN trade t ON t.id=a.trade_id "
        "WHERE vi.project_id=? AND vi.resolved=0 "
        "ORDER BY vi.severity, vi.rule", (project_id,))
    for r in rows:
        w.writerow([r["severity"], r["rule"], r["trade_code"] or "", r["instance_name"] or "",
                    r["field_key"] or "", r["message"], r["expected"]])
    return buf.getvalue()


def _fields_for_export(conn, project_id, include_sensitive):
    fields = [dict(r) for r in conn.execute(
        "SELECT field_key, display_name FROM schema_field "
        "WHERE project_id=? AND visible=1 AND field_key!='instance_name' "
        "ORDER BY export_order", (project_id,))]
    if not include_sensitive:
        fields = [f for f in fields if not _is_sensitive(f["field_key"])]
    return fields


def export_register_xlsx(conn: sqlite3.Connection, project_id: int,
                         trade_id: int | None = None,
                         include_sensitive: bool = False) -> bytes:
    fields = _fields_for_export(conn, project_id, include_sensitive)
    headers = ["Trade", "Instance Name"] + [f["display_name"] for f in fields]
    proj = conn.execute("SELECT name FROM project WHERE id=?", (project_id,)).fetchone()
    wb = Workbook()
    ws = wb.active
    ws.title = "Register"
    start = _style_sheet(ws, headers, f"Asset Register — {proj['name'] if proj else project_id}")
    q = ("SELECT a.instance_name, a.metadata, t.code AS trade_code FROM asset a "
         "JOIN trade t ON t.id=a.trade_id WHERE a.project_id=?")
    params: list = [project_id]
    if trade_id:
        q += " AND a.trade_id=?"
        params.append(trade_id)
    q += " ORDER BY t.code, a.instance_name"
    r = start
    for a in conn.execute(q, params):
        meta = json.loads(a["metadata"] or "{}")
        ws.cell(row=r, column=1, value=a["trade_code"])
        ws.cell(row=r, column=2, value=a["instance_name"])
        for c, f in enumerate(fields, start=3):
            ws.cell(row=r, column=c, value=meta.get(f["field_key"], ""))
        r += 1
    _autofit(ws, len(headers))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_issues_xlsx(conn: sqlite3.Connection, project_id: int) -> bytes:
    headers = ["Severity", "Rule", "Trade", "Instance Name", "Field", "Message", "Expected"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Issues"
    start = _style_sheet(ws, headers, "Validation Issues")
    rows = conn.execute(
        "SELECT vi.severity, vi.rule, vi.field_key, vi.message, vi.expected, "
        "a.instance_name, t.code AS trade_code FROM validation_issue vi "
        "LEFT JOIN asset a ON a.id=vi.asset_id LEFT JOIN trade t ON t.id=a.trade_id "
        "WHERE vi.project_id=? AND vi.resolved=0 ORDER BY vi.severity, vi.rule", (project_id,))
    r = start
    for x in rows:
        for c, v in enumerate([x["severity"], x["rule"], x["trade_code"] or "",
                               x["instance_name"] or "", x["field_key"] or "",
                               x["message"], x["expected"]], start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1
    _autofit(ws, len(headers))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_overview_report_xlsx(conn: sqlite3.Connection, project_id: int, overview: dict) -> bytes:
    """A one-file progress report: summary metrics + per-trade + issues."""
    proj = conn.execute("SELECT name FROM project WHERE id=?", (project_id,)).fetchone()
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=f"Progress Report — {proj['name'] if proj else project_id}").font = _TITLE_FONT
    metrics = [
        ("Total assets", overview["total_assets"]),
        ("Metadata completeness %", overview["metadata_completeness"]),
        ("Assets with errors", overview["assets_with_errors"]),
        ("Cross-trade duplicates", overview["duplicate_instance_names"]),
        ("Naming compliant", overview["naming_compliant"]),
        ("Missing mandatory", overview["missing_mandatory"]),
        ("QR required", overview["qr_required"]),
    ]
    _style_sheet(ws, ["Metric", "Value"], None)
    for i, (k, v) in enumerate(metrics, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    _autofit(ws, 2)
    ws2 = wb.create_sheet("Trades")
    start = _style_sheet(ws2, ["Trade", "Name", "Assets", "Completeness %", "Issues"])
    for i, t in enumerate(overview["trades"], start=start):
        for c, v in enumerate([t["code"], t["name"], t["count"], t["completeness"], t["issues"]], start=1):
            ws2.cell(row=i, column=c, value=v)
    _autofit(ws2, 5)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
