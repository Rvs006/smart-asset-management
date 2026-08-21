"""Import XLSX/CSV into references, schema and assets.

The bundled templates put a title + guidance row above the header, so we locate
the header row heuristically (first row that contains several expected tokens).
User-supplied clean tables (headers in row 1) also work.
"""
from __future__ import annotations

import csv
import io
import json
import re
import sqlite3
from typing import Any

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - openpyxl always installed in the venv
    load_workbook = None

MAX_UPLOAD_BYTES = 20 * 1024 * 1024  # 20 MB cap on any import file


def _guard_size(data: bytes) -> None:
    if len(data) > MAX_UPLOAD_BYTES:
        raise ValueError(f"Upload exceeds the {MAX_UPLOAD_BYTES // (1024 * 1024)} MB limit.")


def normalize_key(header: str) -> str:
    h = str(header or "").strip().lower()
    h = h.replace("&", "and")
    h = re.sub(r"[^\w\s]", "", h)          # drop punctuation (?, /, -)
    h = re.sub(r"\s+", "_", h).strip("_")
    return h


def _rows_from_matrix(matrix: list[list[Any]], expected: list[str]) -> list[dict]:
    exp = {normalize_key(e) for e in expected}
    header_row = 0
    best = -1
    for i, row in enumerate(matrix[:12]):
        keys = {normalize_key(c) for c in row if c not in (None, "")}
        hit = len(keys & exp)
        if hit > best and hit >= min(2, len(exp)):
            best, header_row = hit, i
    headers = [normalize_key(c) for c in matrix[header_row]]
    out = []
    for row in matrix[header_row + 1:]:
        if all(c in (None, "") for c in row):
            continue
        rec = {}
        for j, h in enumerate(headers):
            if h:
                rec[h] = row[j] if j < len(row) else None
        # skip obvious guidance/example filler rows
        if any(v not in (None, "") for v in rec.values()):
            out.append(rec)
    return out


def read_sheet(data: bytes, filename: str, expected: list[str],
               sheet_name: str | None = None) -> list[dict]:
    if filename.lower().endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        matrix = [row for row in csv.reader(io.StringIO(text))]
        return _rows_from_matrix(matrix, expected)
    if load_workbook is None:
        raise RuntimeError("openpyxl not available for XLSX import")
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = next((wb[s] for s in wb.sheetnames if s.upper() != "README"), wb.active)
    matrix = [list(r) for r in ws.iter_rows(values_only=True)]
    return _rows_from_matrix(matrix, expected)


# ---- references -----------------------------------------------------------
REF_COLUMNS = {
    "building": (["Building Name", "Building Reference"], "building_reference", "building_name"),
    "level": (["Level Name", "Level Reference"], "level_reference", "level_name"),
    "space": (["Space ID", "Space Description"], "space_id", "space_description"),
    "system": (["System Code", "System Name"], "system_code", "system_name"),
    "zone": (["Operational Zone Reference", "Zone Name"], "operational_zone_reference", "zone_name"),
    "equipment_type": (["Equipment Type", "BDNS Abbreviation"], "equipment_type", "bdns_abbreviation"),
}


def import_reference_sheet(conn: sqlite3.Connection, project_id: int, kind: str,
                           data: bytes, filename: str) -> int:
    _guard_size(data)
    expected, code_col, label_col = REF_COLUMNS[kind]
    rows = read_sheet(data, filename, expected, sheet_name=None)
    n = 0
    for r in rows:
        code = str(r.get(normalize_key(code_col), "") or "").strip()
        if not code:
            continue
        label = str(r.get(normalize_key(label_col), "") or "").strip()
        abbr = str(r.get("bdns_abbreviation", "") or "").strip()
        conn.execute(
            "INSERT INTO reference_value (project_id,kind,code,label,abbreviation) "
            "VALUES (?,?,?,?,?)", (project_id, kind, code, label, abbr))
        n += 1
    return n


# ---- schema ---------------------------------------------------------------
def import_schema(conn: sqlite3.Connection, project_id: int, data: bytes,
                  filename: str) -> int:
    _guard_size(data)
    expected = ["Field Key", "Display Name", "Data Type", "Required", "Validation Type"]
    rows = read_sheet(data, filename, expected, sheet_name="Field Definitions")
    conn.execute("DELETE FROM schema_field WHERE project_id=?", (project_id,))
    n = 0
    for i, r in enumerate(rows):
        key = str(r.get("field_key", "") or "").strip()
        if not key:
            continue
        conn.execute(
            "INSERT INTO schema_field (project_id,field_key,display_name,grp,data_type,"
            "required,conditional_expr,validation_type,reference_kind,unique_scope,"
            "format_rule,auto_generated,editable,visible,export_order) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (project_id, key, str(r.get("display_name", key)), str(r.get("group", "")),
             str(r.get("data_type", "text")), str(r.get("required", "no")).lower(),
             str(r.get("conditional_required_expression", "") or ""),
             str(r.get("validation_type", "none")).lower(),
             str(r.get("reference_list", "") or "").lower(),
             str(r.get("unique_scope", "") or "").lower(),
             str(r.get("format_rule", "") or ""),
             1 if str(r.get("auto_generated", "")).lower() in ("yes", "true", "1") else 0,
             0 if str(r.get("editable", "yes")).lower() in ("no", "false", "0") else 1,
             0 if str(r.get("visible", "yes")).lower() in ("no", "false", "0") else 1,
             int(r.get("export_order") or i + 1)))
        n += 1
    return n


# ---- assets ---------------------------------------------------------------
def import_assets(conn: sqlite3.Connection, project_id: int, trade_id: int,
                  data: bytes, filename: str, mode: str = "create") -> dict:
    _guard_size(data)
    fields = [dict(r) for r in conn.execute(
        "SELECT field_key, display_name FROM schema_field WHERE project_id=?", (project_id,))]
    keymap = {normalize_key(f["display_name"]): f["field_key"] for f in fields}
    for f in fields:
        keymap[f["field_key"]] = f["field_key"]
    rows = read_sheet(data, filename, [f["display_name"] for f in fields], sheet_name="CONTRACTOR ASSET")
    created = updated = skipped = 0
    errors: list[str] = []
    for r in rows:
        meta: dict[str, Any] = {}
        for raw_key, val in r.items():
            fk = keymap.get(raw_key)
            if fk and val not in (None, ""):
                meta[fk] = val
        instance = str(meta.pop("instance_name", "") or "").strip()
        if not instance and not meta:
            continue
        if mode == "upsert" and instance:
            existing = conn.execute(
                "SELECT id FROM asset WHERE project_id=? AND instance_name=?",
                (project_id, instance)).fetchone()
            if existing:
                conn.execute("UPDATE asset SET metadata=?, trade_id=?, source='import', "
                             "updated_at=datetime('now') WHERE id=?",
                             (json.dumps(meta), trade_id, existing["id"]))
                updated += 1
                continue
        conn.execute("INSERT INTO asset (project_id,trade_id,instance_name,metadata,source) "
                     "VALUES (?,?,?,?, 'import')",
                     (project_id, trade_id, instance, json.dumps(meta)))
        created += 1
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}
